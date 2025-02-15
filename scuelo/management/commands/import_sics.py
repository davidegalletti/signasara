import pandas as pd
from django.core.management.base import BaseCommand
from django.utils import timezone
from pathlib import Path
from scuelo.models import AnneeScolaire, Eleve, Inscription, Classe, Ecole, TypeClasse, CONDITION_ELEVE, SEX, CS_PY, HAND
from cash.models import Mouvement, Tarif
import logging
from datetime import datetime
from openpyxl import load_workbook

# Define log file paths
LOG_DIR = "logs"
ECOLE_LOG_FILE = f"logs/import_ecole.log"
TYPE_CLASSE_LOG_FILE = f"logs/import_type_classe.log"
CLASSE_LOG_FILE = f"logs/import_classe.log"
ELEVE_LOG_FILE = f"logs/import_eleve.log"
PAIEMENT_LOG_FILE = f"logs/import_paiement.log"


def setup_logger(name, log_file, level=logging.INFO):
    """To setup as many loggers as you want"""

    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

    handler = logging.FileHandler(log_file)
    handler.setFormatter(formatter)

    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)

    return logger

# Setup loggers
ecole_logger = setup_logger('ecole_importer', ECOLE_LOG_FILE)
type_classe_logger = setup_logger('type_classe_importer', TYPE_CLASSE_LOG_FILE)
classe_logger = setup_logger('classe_importer', CLASSE_LOG_FILE)
eleve_logger = setup_logger('eleve_importer', ELEVE_LOG_FILE)
paiement_logger = setup_logger('paiement_importer', PAIEMENT_LOG_FILE)

def parse_date(date_str):
    if isinstance(date_str, datetime):
        return date_str.date()

    if isinstance(date_str, str):
        date_str = date_str.strip().replace('\xa0', '')
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            try:
                return datetime.strptime(date_str, '%d/%m/%Y').date()
            except ValueError:
                return None
    return None

class Command(BaseCommand):
    help = 'Imports data from SICS Excel file, including Ecole, Classe, Eleve, and Paiement with separate logs.'

    def handle(self, *args, **options):
        BASE_DIR = str(Path(__file__).resolve().parent.parent.parent)
        file_path = f'{BASE_DIR}/export_sics/combined_excel.xlsx'  # Use the combined Excel file
        failed_eleves = []

        # Create logs directory if it doesn't exist
        Path(LOG_DIR).mkdir(parents=True, exist_ok=True)

        logger = logging.getLogger(__name__)
        logger.info(f'Starting import from {file_path}')

        try:
            # Load workbook using openpyxl
            wb = load_workbook(file_path)

            # Load worksheets
            ws_eleve = wb['BcK Studente']
            ws_paiement = wb['BcK Pagamento']

        except FileNotFoundError:
            logger.error(f"File not found: {file_path}")
            return
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            return

        annee_scolaire_actuel = AnneeScolaire.objects.get(actuel=True)
        derniere_annee_scolaire = AnneeScolaire.objects.filter(actuel=False).order_by('-date_initiale').first()

        if not derniere_annee_scolaire:
            logger.error("No previous school year found.")
            return

        # 1. Manually Create Ecole
        ecole_logger.info("Creating Ecoles...")
        ecole_data = [
            {"nom": "Sc_TBD", "ville": "TBD"},
            {"nom": "Sc_Nas_Mat", "ville": "Nas_Mat"},
            {"nom": "Sc_Nas_Pri", "ville": "Nas_Pri"},
            {"nom": "Sc_LMS", "ville": "LMS"},
            {"nom": "Sc_WLM", "ville": "WLM"},
            {"nom": "Sc_LPMSBD", "ville": "LPMSBD"},
            {"nom": "Sc_WP", "ville": "WP"},
            {"nom": "Sc_EPCELP", "ville": "EPCELP"},
            {"nom": "Sc_EPPLE", "ville": "EPPLE"},
            {"nom": "Sc_YAMT", "ville": "YAMT"},
            {"nom": "Sc_LSB", "ville": "LSB"},
        ]

        for data in ecole_data:
            try:
                ecole, created = Ecole.objects.get_or_create(
                    nom=data["nom"],
                    ville=data["ville"],
                    defaults={
                        'nom_du_referent': 'N/A',
                        'prenom_du_referent': 'N/A',
                        'email_du_referent': 'N/A',
                        'telephone_du_referent': 'N/A',
                        'note': 'N/A',
                        'externe': True
                    }
                )
                if created:
                    ecole_logger.info(f"Created Ecole: {ecole.nom}")
                else:
                    ecole_logger.info(f"Ecole already exists: {ecole.nom}")
            except Exception as e:
                ecole_logger.error(f"Error creating Ecole {data['nom']}: {e}")

        # 2. Manually Create TypeClasse
        type_classe_logger.info("Creating TypeClasses...")
        type_classe_data = [
            {"nom": "PS", "ordre": 1, "type_ecole": "M", "legacy_id": "PK_PS"},
            {"nom": "MS", "ordre": 2, "type_ecole": "M", "legacy_id": "PK_MS"},
            {"nom": "GS", "ordre": 3, "type_ecole": "M", "legacy_id": "PK_GS"},
            {"nom": "CP1", "ordre": 4, "type_ecole": "P", "legacy_id": "PK_CP1"},
            {"nom": "CP2", "ordre": 5, "type_ecole": "P", "legacy_id": "PK_CP2"},
            {"nom": "CE1", "ordre": 6, "type_ecole": "P", "legacy_id": "PK_CE1"},
            {"nom": "CE2", "ordre": 7, "type_ecole": "P", "legacy_id": "PK_CE2"},
            {"nom": "CM1", "ordre": 8, "type_ecole": "P", "legacy_id": "PK_CM1"},
            {"nom": "CM2", "ordre": 9, "type_ecole": "P", "legacy_id": "PK_CM2"},
            {"nom": "6me", "ordre": 10, "type_ecole": "L", "legacy_id": "PK_6me"},
            {"nom": "5me", "ordre": 11, "type_ecole": "L", "legacy_id": "PK_5me"},
            {"nom": "4me", "ordre": 12, "type_ecole": "L", "legacy_id": "PK_4me"},
            {"nom": "3me", "ordre": 13, "type_ecole": "L", "legacy_id": "PK_3me"},
            {"nom": "2me", "ordre": 14, "type_ecole": "L", "legacy_id": "PK_2me"},
            {"nom": "1ere", "ordre": 15, "type_ecole": "L", "legacy_id": "PK_1ere"},
            {"nom": "Term", "ordre": 16, "type_ecole": "L", "legacy_id": "PK_Term"},
        ]

        for data in type_classe_data:
            try:
                type_classe, created = TypeClasse.objects.get_or_create(
                    nom=data["nom"],
                    type_ecole=data["type_ecole"],
                    defaults={"ordre": data["ordre"]}
                )
                if created:
                    type_classe_logger.info(f"Created TypeClasse: {type_classe.nom}")
                else:
                    type_classe_logger.info(f"TypeClasse already exists: {type_classe.nom}")
            except Exception as e:
                type_classe_logger.error(f"Error creating TypeClasse {data['nom']}: {e}")

        # 3. Manually Create Classe
        classe_logger.info("Creating Classes...")
        classe_data = [
            {"ecole_nom": "Sc_TBD", "type_classe_nom": "6me", "nom": "6me-TBD", "legacy_id": "_PK-6me-TBD"},
            {"ecole_nom": "Sc_TBD", "type_classe_nom": "5me", "nom": "5me-TBD", "legacy_id": "_PK-5me-TBD"},
            {"ecole_nom": "Sc_TBD", "type_classe_nom": "4me", "nom": "4me-TBD", "legacy_id": "_PK-4me-TBD"},
            {"ecole_nom": "Sc_TBD", "type_classe_nom": "3me", "nom": "3me-TBD", "legacy_id": "_PK-3me-TBD"},
            {"ecole_nom": "Sc_TBD", "type_classe_nom": "2me", "nom": "2me-TBD", "legacy_id": "_PK-2me-TBD"},
            {"ecole_nom": "Sc_TBD", "type_classe_nom": "1ere", "nom": "1ere-TBD", "legacy_id": "_PK-1ere-TBD"},
            {"ecole_nom": "Sc_TBD", "type_classe_nom": "Term", "nom": "Term-TBD", "legacy_id": "_PK-Term-TBD"},
            {"ecole_nom": "Sc_Nas_Mat", "type_classe_nom": "PS", "nom": "PS-Nas_Mat", "legacy_id": "_PK-PS-Nas"},
            {"ecole_nom": "Sc_Nas_Mat", "type_classe_nom": "MS", "nom": "MS-Nas_Mat", "legacy_id": "_PK-MS-Nas"},
            {"ecole_nom": "Sc_Nas_Mat", "type_classe_nom": "GS", "nom": "GS-Nas_Mat", "legacy_id": "_PK-GS-Nas"},
            {"ecole_nom": "Sc_Nas_Pri", "type_classe_nom": "CP1", "nom": "CP1-Nas_Pri", "legacy_id": "_PK-CP1-Nas"},
            {"ecole_nom": "Sc_Nas_Pri", "type_classe_nom": "CP2", "nom": "CP2-Nas_Pri", "legacy_id": "_PK-CP2-Nas"},
            {"ecole_nom": "Sc_Nas_Pri", "type_classe_nom": "CE1", "nom": "CE1-Nas_Pri", "legacy_id": "_PK-CE1-Nas"},
            {"ecole_nom": "Sc_Nas_Pri", "type_classe_nom": "CE2", "nom": "CE2-Nas_Pri", "legacy_id": "_PK-CE2-Nas"},
            {"ecole_nom": "Sc_Nas_Pri", "type_classe_nom": "CM1", "nom": "CM1-Nas_Pri", "legacy_id": "_PK-CM1-Nas"},
            {"ecole_nom": "Sc_Nas_Pri", "type_classe_nom": "CM2", "nom": "CM2-Nas_Pri", "legacy_id": "_PK-CM2-Nas"},
            {"ecole_nom": "Sc_LMS", "type_classe_nom": "6me", "nom": "6me-LMS", "legacy_id": "_PK-6me-LMS"},
            {"ecole_nom": "Sc_LMS", "type_classe_nom": "5me", "nom": "5me-LMS", "legacy_id": "_PK-5me-LMS"},
            {"ecole_nom": "Sc_LMS", "type_classe_nom": "4me", "nom": "4me-LMS", "legacy_id": "_PK-4me-LMS"},
            {"ecole_nom": "Sc_LMS", "type_classe_nom": "3me", "nom": "3me-LMS", "legacy_id": "_PK-3me-LMS"},
            {"ecole_nom": "Sc_LMS", "type_classe_nom": "2me", "nom": "2me-LMS", "legacy_id": "_PK-2me-LMS"},
            {"ecole_nom": "Sc_LMS", "type_classe_nom": "1ere", "nom": "1ere-LMS", "legacy_id": "_PK-1ere-LMS"},
            {"ecole_nom": "Sc_LMS", "type_classe_nom": "Term", "nom": "Term-LMS", "legacy_id": "_PK-Term-LMS"},
            {"ecole_nom": "Sc_WLM", "type_classe_nom": "6me", "nom": "6me-WLM", "legacy_id": "_PK-6me-WLM"},
            {"ecole_nom": "Sc_WLM", "type_classe_nom": "5me", "nom": "5me-WLM", "legacy_id": "_PK-5me-WLM"},
            {"ecole_nom": "Sc_WLM", "type_classe_nom": "4me", "nom": "4me-WLM", "legacy_id": "_PK-4me-WLM"},
            {"ecole_nom": "Sc_WLM", "type_classe_nom": "3me", "nom": "3me-WLM", "legacy_id": "_PK-3me-WLM"},
            {"ecole_nom": "Sc_WLM", "type_classe_nom": "2me", "nom": "2me-WLM", "legacy_id": "_PK-2me-WLM"},
            {"ecole_nom": "Sc_WLM", "type_classe_nom": "1ere", "nom": "1ere-WLM", "legacy_id": "_PK-1ere-WLM"},
            {"ecole_nom": "Sc_WLM", "type_classe_nom": "Term", "nom": "Term-WLM", "legacy_id": "_PK-Term-WLM"},
            {"ecole_nom": "Sc_LPMSBD", "type_classe_nom": "6me", "nom": "6me-LPMSBD", "legacy_id": "_PK-6me-LPMSBD"},
            {"ecole_nom": "Sc_LPMSBD", "type_classe_nom": "5me", "nom": "5me-LPMSBD", "legacy_id": "_PK-5me-LPMSBD"},
            {"ecole_nom": "Sc_LPMSBD", "type_classe_nom": "4me", "nom": "4me-LPMSBD", "legacy_id": "_PK-4me-LPMSBD"},
            {"ecole_nom": "Sc_LPMSBD", "type_classe_nom": "3me", "nom": "3me-LPMSBD", "legacy_id": "_PK-3me-LPMSBD"},
            {"ecole_nom": "Sc_LPMSBD", "type_classe_nom": "2me", "nom": "2me-LPMSBD", "legacy_id": "_PK-2me-LPMSBD"},
            {"ecole_nom": "Sc_LPMSBD", "type_classe_nom": "1ere", "nom": "1ere-LPMSBD", "legacy_id": "_PK-1ere-LPMSBD"},
            {"ecole_nom": "Sc_LPMSBD", "type_classe_nom": "Term", "nom": "Term-LPMSBD", "legacy_id": "_PK-Term-LPMSBD"},
             {"ecole_nom": "Sc_WP", "type_classe_nom": "6me", "nom": "6me-WP", "legacy_id": "_PK-6me-WP"},
            {"ecole_nom": "Sc_WP", "type_classe_nom": "5me", "nom": "5me-WP", "legacy_id": "_PK-5me-WP"},
            {"ecole_nom": "Sc_WP", "type_classe_nom": "4me", "nom": "4me-WP", "legacy_id": "_PK-4me-WP"},
            {"ecole_nom": "Sc_WP", "type_classe_nom": "3me", "nom": "3me-WP", "legacy_id": "_PK-3me-WP"},
            {"ecole_nom": "Sc_WP", "type_classe_nom": "2me", "nom": "2me-WP", "legacy_id": "_PK-2me-WP"},
            {"ecole_nom": "Sc_WP", "type_classe_nom": "1ere", "nom": "1ere-WP", "legacy_id": "_PK-1ere-WP"},
            {"ecole_nom": "Sc_WP", "type_classe_nom": "Term", "nom": "Term-WP", "legacy_id": "_PK-Term-WP"},
            {"ecole_nom": "Sc_EPCELP", "type_classe_nom": "CP1", "nom": "CP1-EPCELP", "legacy_id": "_PK-CP1-EPCELP"},
            {"ecole_nom": "Sc_EPCELP", "type_classe_nom": "CP2", "nom": "CP2-EPCELP", "legacy_id": "_PK-CP2-EPCELP"},
            {"ecole_nom": "Sc_EPCELP", "type_classe_nom": "CE1", "nom": "CE1-EPCELP", "legacy_id": "_PK-CE1-EPCELP"},
            {"ecole_nom": "Sc_EPCELP", "type_classe_nom": "CE2", "nom": "CE2-EPCELP", "legacy_id": "_PK-CE2-EPCELP"},
            {"ecole_nom": "Sc_EPCELP", "type_classe_nom": "CM1", "nom": "CM1-EPCELP", "legacy_id": "_PK-CM1-EPCELP"},
            {"ecole_nom": "Sc_EPCELP", "type_classe_nom": "CM2", "nom": "CM2-EPCELP", "legacy_id": "_PK-CM2-EPCELP"},
            {"ecole_nom": "Sc_EPPLE", "type_classe_nom": "CP1", "nom": "CP1-EPPLE", "legacy_id": "_PK-CP1-EPPLE"},
            {"ecole_nom": "Sc_EPPLE", "type_classe_nom": "CP2", "nom": "CP2-EPPLE", "legacy_id": "_PK-CP2-EPPLE"},
            {"ecole_nom": "Sc_EPPLE", "type_classe_nom": "CE1", "nom": "CE1-EPPLE", "legacy_id": "_PK-CE1-EPPLE"},
            {"ecole_nom": "Sc_EPPLE", "type_classe_nom": "CE2", "nom": "CE2-EPPLE", "legacy_id": "_PK-CE2-EPPLE"},
            {"ecole_nom": "Sc_EPPLE", "type_classe_nom": "CM1", "nom": "CM1-EPPLE", "legacy_id": "_PK-CM1-EPPLE"},
            {"ecole_nom": "Sc_EPPLE", "type_classe_nom": "CM2", "nom": "CM2-EPPLE", "legacy_id": "_PK-CM2-EPPLE"},
            {"ecole_nom": "Sc_YAMT", "type_classe_nom": "CP1", "nom": "CP1-YAMT", "legacy_id": "_PK-CP1-YAMT"},
            {"ecole_nom": "Sc_YAMT", "type_classe_nom": "CP2", "nom": "CP2-YAMT", "legacy_id": "_PK-CP2-YAMT"},
            {"ecole_nom": "Sc_YAMT", "type_classe_nom": "CE1", "nom": "CE1-YAMT", "legacy_id": "_PK-CE1-YAMT"},
            {"ecole_nom": "Sc_YAMT", "type_classe_nom": "CE2", "nom": "CE2-YAMT", "legacy_id": "_PK-CE2-YAMT"},
            {"ecole_nom": "Sc_YAMT", "type_classe_nom": "CM1", "nom": "CM1-YAMT", "legacy_id": "_PK-CM1-YAMT"},
            {"ecole_nom": "Sc_YAMT", "type_classe_nom": "CM2", "nom": "CM2-YAMT", "legacy_id": "_PK-CM2-YAMT"},
             {"ecole_nom": "Sc_LSB", "type_classe_nom": "CP1", "nom": "CP1-LSB", "legacy_id": "_PK-CP1-LSB"},
            {"ecole_nom": "Sc_LSB", "type_classe_nom": "CP2", "nom": "CP2-LSB", "legacy_id": "_PK-CP2-LSB"},
            {"ecole_nom": "Sc_LSB", "type_classe_nom": "CE1", "nom": "CE1-LSB", "legacy_id": "_PK-CE1-LSB"},
            {"ecole_nom": "Sc_LSB", "type_classe_nom": "CE2", "nom": "CE2-LSB", "legacy_id": "_PK-CE2-LSB"},
            {"ecole_nom": "Sc_LSB", "type_classe_nom": "CM1", "nom": "CM1-LSB", "legacy_id": "_PK-CM1-LSB"},
            {"ecole_nom": "Sc_LSB", "type_classe_nom": "CM2", "nom": "CM2-LSB", "legacy_id": "_PK-CM2-LSB"},
        ]

        for data in classe_data:
            try:
                # Retrieve Ecole and TypeClasse instances
                ecole = Ecole.objects.get(nom=data["ecole_nom"])
                type_classe = TypeClasse.objects.get(nom=data["type_classe_nom"])

                classe, created = Classe.objects.get_or_create(
                    ecole=ecole,
                    type=type_classe,
                    legacy_id=data["legacy_id"],
                    nom=data["nom"],
                )

                if created:
                    classe_logger.info(f"Created Classe: {classe.nom}")
                else:
                    classe_logger.info(f"Classe already exists: {classe.nom}")

            except Ecole.DoesNotExist:
                classe_logger.error(f"Ecole with nom {data['ecole_nom']} not found.")
                continue
            except TypeClasse.DoesNotExist:
                classe_logger.error(f"TypeClasse with nom {data['type_classe_nom']} not found.")
                continue
            except Exception as e:
                classe_logger.error(f"Error creating Classe {data['nom']}: {e}")

        # 4. Import Eleve
        eleve_logger.info("Importing Eleves...")
        columns_eleve = {}
        ws_eleve = wb['BcK Studente']
        failed_eleves_data = [] # List to store dictionaries of failed eleves

        for row in ws_eleve.iter_rows(min_row=1):
            row_values = [cell.value for cell in row]  # Extract values manually

            if len(columns_eleve.keys()) == 0:
                for column_index, column_name in enumerate(row_values):
                    columns_eleve[column_name] = column_index
            else:
                try:
                    classe_legacy_id = row_values[columns_eleve['__FK_Classe_ID']]
                    eleve_id = row_values[columns_eleve['_PK_Studente_ID']]

                    # Extract student data
                    student_data = {
                        "classe_legacy_id": classe_legacy_id,
                        "eleve_id": eleve_id,
                        "condition_eleve_value": row_values[columns_eleve['Condition_eleve']],
                        "sex_value": row_values[columns_eleve['Sex']],
                        "cs_py_value": row_values[columns_eleve['CS_PY']],
                        "hand_value": row_values[columns_eleve['Hand']],
                        "D_enquete": row_values[columns_eleve['D_enquete']],
                        "Nom": row_values[columns_eleve['Nom']],
                        "Prenom": row_values[columns_eleve['Prenom']],
                        "Date-Naissance": row_values[columns_eleve['Date-Naissance']],
                        "A_inscr": row_values[columns_eleve['A_inscr']],
                        "Parent": row_values[columns_eleve['Parent']],
                        "Tel_parent": row_values[columns_eleve['Tel_parent']]
                    }
                     #here transform cs_py_value
                    condition_eleve_value = student_data["condition_eleve_value"]
                    sex_value = student_data["sex_value"]
                    cs_py_value = student_data["cs_py_value"]
                    hand_value = student_data["hand_value"]

                    #here i do the data validation
                    if str(condition_eleve_value).strip() not in dict(CONDITION_ELEVE).keys():
                        eleve_logger.error(f"Invalid CONDITION_ELEVE value: {condition_eleve_value} in row")
                        student_data['error'] = f"Invalid CONDITION_ELEVE value: {condition_eleve_value}"  # Add the error in data
                        failed_eleves_data.append(student_data)
                        continue
                     # Validating sex and mapping if it is None  
                    if str(sex_value).strip() not in dict(SEX).keys():
                        eleve_logger.error(f"Invalid SEX value: {sex_value} in row")
                        student_data['error'] = f"Invalid SEX value: {sex_value}"
                        failed_eleves_data.append(student_data)
                        continue

                    cs_py_mapping = {
                            'CS': 'C',
                            'EXTRA': 'E',
                            'PY': 'P',
                            'ACC': 'A',
                            'BRAVO': 'B',
                        }
                    # Mapping
                    cs_py_value = cs_py_mapping.get(str(cs_py_value).strip(), None)

                    if cs_py_value is None:
                        eleve_logger.error(f"Invalid CS_PY value: {cs_py_value} in row. Valid values are: {', '.join(dict(CS_PY).keys())}")
                        student_data['error'] = f"Invalid CS_PY value: {cs_py_value}"
                        failed_eleves_data.append(student_data)
                        continue
                    
                    hand_value = str(hand_value).strip() if hand_value else None

                    if hand_value and hand_value not in dict(HAND).keys():
                        eleve_logger.error(f"Invalid Hand value: {hand_value} in row")
                        student_data['error'] = f"Invalid Hand value: {hand_value}"
                        failed_eleves_data.append(student_data)
                        continue
                    if Classe.objects.filter(legacy_id=classe_legacy_id).exists():
                        classe = Classe.objects.get(legacy_id=classe_legacy_id)
                        new_e, created = Eleve.objects.get_or_create(
                        pk=eleve_id,
                        defaults={
                            'legacy_id': eleve_id,
                            'date_enquete': parse_date(student_data['D_enquete']),
                            'nom': student_data['Nom'],
                            'prenom': student_data['Prenom'],
                            'condition_eleve': condition_eleve_value,
                            'sex': sex_value,
                            'date_naissance': parse_date(student_data['Date-Naissance']),
                            'cs_py': cs_py_value,
                            'hand': hand_value or None,  # Set to None if hand_value is blank,
                            'annee_inscr': student_data['A_inscr'],
                            'parent': student_data['Parent'],
                            'tel_parent': student_data['Tel_parent']
                        }
                        )
                        # Create inscription
                        inscription, inscription_created = Inscription.objects.get_or_create(
                            eleve=new_e,
                            annee_scolaire=derniere_annee_scolaire,
                            classe=classe
                        )
                        if created:
                            eleve_logger.info(f"Student {new_e.nom} {new_e.prenom} imported successfully.")
                        else:
                            eleve_logger.info(f"Eleve {eleve_id} already exists.")
                    else:
                        eleve_logger.error(f"Class with legacy_id {classe_legacy_id} not found for student {student_data['Nom']}.")
                        student_data['error'] = f"Class with legacy_id {classe_legacy_id} not found"
                        failed_eleves_data.append(student_data)
                except Exception as ex:
                    eleve_logger.error(f"Error importing student {student_data['Nom']}: {str(ex)}")
                    student_data['error'] = str(ex)
                    failed_eleves_data.append(student_data)

        # 5. Create Excel file for failed eleves
        if failed_eleves_data:
            failed_eleves_df = pd.DataFrame(failed_eleves_data)
            failed_eleves_file = f'{LOG_DIR}/failed_eleves.xlsx'
            failed_eleves_df.to_excel(failed_eleves_file, index=False)
            logger.warning(f"Failed eleves data saved to {failed_eleves_file}")

        '''  # 6. Process Paiement
        paiement_logger.info("Importing Paiements...")
        columns_paiement = {}
        ws_paiement = wb['BcK Pagamento']
        for row in ws_paiement.iter_rows(min_row=1):
            row_values = [cell.value for cell in row]  # Extract values manually

            if len(columns_paiement.keys()) == 0:
                for column_index, column_name in enumerate(row_values):
                    columns_paiement[column_name] = column_index
            else:
                try:
                    eleve_legacy_id = row_values[columns_paiement['__FK_Studente']]
                    try:
                        eleve = Eleve.objects.get(pk=eleve_legacy_id)
                        inscription = Inscription.objects.get(annee_scolaire=derniere_annee_scolaire, eleve=eleve)

                        new_p = Mouvement(
                            legacy_id=row_values[columns_paiement['_PK_Pagamento_ID']],
                            causal=row_values[columns_paiement['Causale_Pagamento']],
                            montant=row_values[columns_paiement['Importo_Pagamento']],
                            date_paye=parse_date(row_values[columns_paiement['D_pagamento']]),
                            note=row_values[columns_paiement['Note_Pagamento']],
                            inscription=inscription
                        )
                        new_p.save()
                        paiement_logger.info(f"Payment for {eleve.nom} {eleve.prenom} imported successfully.")
                    except Eleve.DoesNotExist:
                        paiement_logger.error(f"Eleve with ID {eleve_legacy_id} does not exist.")
                    except Inscription.DoesNotExist:
                        paiement_logger.error(f"Inscription for Eleve ID {eleve_legacy_id} in {derniere_annee_scolaire.nom} not found.")
                except Exception as ex:
                    paiement_logger.error(f"Error importing payment for Eleve ID {row_values[columns_paiement['__FK_Studente']]}: {str(ex)}")

        logger.info('Import SICS END')
        '''
        # 6. Process Paiement
        paiement_logger.info("Importing Paiements...")
        columns_paiement = {}
        ws_paiement = wb['BcK Pagamento']
        for row in ws_paiement.iter_rows(min_row=1):
            row_values = [cell.value for cell in row]  # Extract values manually

            if len(columns_paiement.keys()) == 0:
                for column_index, column_name in enumerate(row_values):
                    columns_paiement[column_name] = column_index
            else:
                try:
                    eleve_legacy_id = row_values[columns_paiement['__FK_Studente']]
                    try:
                        eleve = Eleve.objects.get(pk=eleve_legacy_id)
                        inscription = Inscription.objects.get(annee_scolaire=derniere_annee_scolaire, eleve=eleve)

                        new_p = Mouvement(
                            causal=row_values[columns_paiement['Causale_Pagamento']],
                            montant=row_values[columns_paiement['Importo_Pagamento']],
                            date_paye=parse_date(row_values[columns_paiement['D_pagamento']]),
                            note=row_values[columns_paiement['Note_Pagamento']],
                            inscription=inscription
                        )
                        new_p.save()
                        paiement_logger.info(f"Payment for {eleve.nom} {eleve.prenom} imported successfully.")
                    except Eleve.DoesNotExist:
                        paiement_logger.error(f"Eleve with ID {eleve_legacy_id} does not exist.")
                    except Inscription.DoesNotExist:
                        paiement_logger.error(f"Inscription for Eleve ID {eleve_legacy_id} in {derniere_annee_scolaire.nom} not found.")
                except Exception as ex:
                    paiement_logger.error(f"Error importing payment for Eleve ID {row_values[columns_paiement['__FK_Studente']]}: {str(ex)}")



        logger.info('Import SICS END')