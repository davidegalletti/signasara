import logging
import uuid
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from pathlib import Path
from cash.models import Expense, Cashier
from datetime import datetime

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("import_expenses.log"), logging.StreamHandler()],
)

logger = logging.getLogger(__name__)

def parse_date(date_str):
    """Convert date string to a date object."""
    if isinstance(date_str, datetime):
        return date_str.date()

    if isinstance(date_str, str):
        date_str = date_str.strip().replace("\xa0", "")
        if not date_str:
            return None

        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):  # Try different formats
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
    return None

class Command(BaseCommand):
    help = "Imports expenses from Excel file."

    def handle(self, *args, **options):
        BASE_DIR = str(Path(__file__).resolve().parent.parent.parent)
        full_path = f"{BASE_DIR}/exportcash/expenses sics.xlsx"  # Update with actual filename

        logger.info(f"Starting expense import from {full_path}")

        try:
            wb = load_workbook(full_path)
        except FileNotFoundError:
            logger.error(f"File not found: {full_path}")
            return

        ws_expense = wb.active  # Assuming data is in the first sheet

        # Ensure C_SCO cashier exists
        cashier_sco, _ = Cashier.objects.get_or_create(name="C_SCO")

        logger.info("Cashier ensured: C_SCO")

        # Read column headers
        headers = [cell.value for cell in ws_expense[1]]
        columns = {name: index for index, name in enumerate(headers)}

        required_columns = ["_PK_Uscita_cassa", "D_uscita", "Importo", "Descrizione"]
        for col in required_columns:
            if col not in columns:
                logger.error(f"Missing required column: {col}")
                return

        for row in ws_expense.iter_rows(min_row=2, values_only=True):
            try:
                legacy_id = row[columns["_PK_Uscita_cassa"]]
                date = parse_date(row[columns["D_uscita"]])
                amount = row[columns["Importo"]]
                description = row[columns["Descrizione"]]

                if not all([legacy_id, date, amount, description]):
                    logger.warning(f"Skipping incomplete row: {row}")
                    continue

                # Ensure legacy_id is stored separately, avoiding ID conflict
                expense, created = Expense.objects.get_or_create(
                    legacy_id=legacy_id,
                    defaults={
                        "date": date,
                        "amount": amount,
                        "description": description,
                        "cashier": cashier_sco,
                    },
                )

                if created:
                    logger.info(f"Expense {legacy_id} imported successfully.")
                else:
                    logger.warning(f"Expense {legacy_id} already exists. Skipping.")

            except Exception as ex:
                logger.error(f"Error processing row: {row} - {ex}")

        logger.info("Expense import completed.")
